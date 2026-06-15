#!/usr/bin/env python3
"""
Meeting Street Time Tracker - Excel Generation Server
Run: python3 xlsx_server.py
Then open the HTML app and export Excel normally.
"""
import http.server, json, io, base64, datetime, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImg

PORT = 8766

# ── Colors ──────────────────────────────────────────
TEAL   = "008080"   # Header row background (matches screenshot)
NAVY   = "05294B"   # Title text
WHITE  = "FFFFFF"
ALT1   = "D5E8F0"   # Alternating row blue
ALT2   = "FFFFFF"   # Alternating row white
GREEN  = "A9DFBF"   # Positive hours remaining
RED    = "F5B7B1"   # Negative hours remaining
STATUS_CLOSED = "A9DFBF"  # Green for Closed
STATUS_OPEN   = "F5B7B1"  # Red/pink for Open
AMBER  = "FAD7A0"   # Amber for other statuses
BLACK  = "000000"

def s_font(bold=False, sz=10, color=BLACK, italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=color, italic=italic)

def s_fill(color):
    return PatternFill("solid", fgColor=color)

def s_align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def s_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, value):
    cell.value = value
    cell.font = s_font(bold=True, sz=10, color=WHITE)
    cell.fill = s_fill(TEAL)
    cell.alignment = s_align("center", "center", True)
    cell.border = s_border(WHITE)

def apply_data(cell, value, row_idx, align="left", num_fmt=None, bold=False, fill_override=None):
    cell.value = value
    if fill_override:
        fg = fill_override
    else:
        fg = ALT1 if row_idx % 2 == 0 else ALT2
    cell.fill = s_fill(fg)
    cell.font = s_font(bold=bold, sz=10)
    cell.alignment = s_align(align, "center", True)
    cell.border = s_border()
    if num_fmt:
        cell.number_format = num_fmt

def get_status_fill(status):
    """Return color fill based on status value."""
    return None

def get_hrs_remaining_fill(hrs):
    """Green if positive, Red if negative."""
    if hrs < 0:
        return RED
    return GREEN

def generate_workbook(payload):
    fp_b64 = payload.get("logo_b64", "")
    months_data = payload.get("months", [])

    wb = Workbook()
    wb.remove(wb.active)

    for md in months_data:
        month    = md["month"]
        allotment= float(md.get("allotment", 30))
        entries  = md.get("entries", [])
        total_billed = sum(float(e.get("billedHours", 0) or 0) for e in entries)
        remaining = allotment - total_billed

        ws = wb.create_sheet(title=month)

        # Column widths matching screenshot template
        for col, width in zip("ABCDEFGHIJKL", [18, 12, 11, 11, 22, 14, 24, 17, 14, 16, 16, 55]):
            ws.column_dimensions[col].width = width

        # Row heights
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 26
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 10
        ws.row_dimensions[5].height = 32

        # ── Logo (A1:C4) ──────────────────────────────────
        if fp_b64:
            try:
                img = XLImg(io.BytesIO(base64.b64decode(fp_b64)))
                img.width = 95; img.height = 80
                img.anchor = "A1"
                ws.add_image(img)
            except: pass

        # ── Title row 2 D–H ───────────────────────────────
        ws.merge_cells("D2:H2")
        tc = ws["D2"]
        tc.value = "CLAFLIN INTERVAL 2026"
        tc.font  = Font(name="Calibri", bold=True, size=16, color=NAVY)
        tc.alignment = s_align("center", "center")

        # ── Stats I1:J3 ────────────────────────────────────
        for r, lbl, val, is_remaining in [
            (1, "Monthly Allotment:", allotment, False),
            (2, "TOTAL HOURS USED:",  total_billed, False),
            (3, "REMAINING:",         remaining,   True),
        ]:
            lc = ws.cell(row=r, column=9)
            lc.value = lbl
            lc.font  = Font(name="Calibri", bold=True, size=10, color=NAVY)
            lc.alignment = s_align("right", "center")
            vc = ws.cell(row=r, column=10)
            vc.value = val
            vc.font  = Font(name="Calibri", bold=True, size=11)
            vc.alignment = s_align("center", "center")
            vc.number_format = "0.00"
            if is_remaining:
                vc.fill = s_fill(get_hrs_remaining_fill(val))

        # ── Column headers row 5 ──────────────────────────
        HDRS = ["Project","Date","Check In","Check Out","Staff Attended",
                "Number of Staff","Client Contact","Session Duration",
                "Billed Hours","Hours Remaining","Status","Work Details/Notes"]
        for i, h in enumerate(HDRS, 1):
            apply_header(ws.cell(row=5, column=i), h)

        # ── Data rows starting row 6 ──────────────────────
        for ridx, e in enumerate(entries):
            rn = 6 + ridx

            dur    = float(e.get("duration", 0) or 0)
            billed = float(e.get("billedHours", 0) or 0)
            hrs_rem= float(e.get("hoursRemaining", allotment) or allotment)
            sc     = int(e.get("numStaff", 0) or 0)
            status = e.get("status", "")

            date_val = e.get("date", "")
            if date_val:
                try:
                    y,m,d = date_val.split("-")
                    date_val = datetime.date(int(y),int(m),int(d))
                except: pass

            apply_data(ws.cell(rn,1), e.get("project",""),   ridx, "center")
            dv = ws.cell(rn,2); dv.value = date_val
            dv.fill = s_fill(ALT1 if ridx%2==0 else ALT2)
            dv.font = s_font(); dv.alignment = s_align("center","center")
            dv.border = s_border(); dv.number_format = "M/D/YYYY"
            apply_data(ws.cell(rn,3), e.get("checkIn",""),    ridx, "center")
            apply_data(ws.cell(rn,4), e.get("checkOut",""),   ridx, "center")
            apply_data(ws.cell(rn,5), e.get("staff",""),      ridx, "center")
            apply_data(ws.cell(rn,6), sc,                     ridx, "center", "0")
            apply_data(ws.cell(rn,7), e.get("client",""),     ridx, "center")
            apply_data(ws.cell(rn,8), dur,                    ridx, "center", "0.00")
            apply_data(ws.cell(rn,9), billed,                 ridx, "center", "0.00")
            # Hours Remaining — conditional color
            apply_data(ws.cell(rn,10), hrs_rem, ridx, "center", "0.00", bold=True,
                       fill_override=get_hrs_remaining_fill(hrs_rem))
            # Status — color-coded
            apply_data(ws.cell(rn,11), status, ridx, "center",
                       fill_override=get_status_fill(status))
            apply_data(ws.cell(rn,12), e.get("notes",""),     ridx, "left")

        # ── Auto-filter on header row ─────────────────────
        ws.auto_filter.ref = f"A5:L{5 + len(entries)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass  # quiet

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin","*")
        self.send_header("Access-Control-Allow-Methods","POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers","Content-Type")
        self.end_headers()

    def do_POST(self):
        if self.path == "/generate":
            length = int(self.headers.get("Content-Length",0))
            body   = self.rfile.read(length)
            try:
                payload  = json.loads(body)
                xlsx_bytes = generate_workbook(payload)
                fn = payload.get("filename", "export.xlsx")
                if not fn.lower().endswith(".xlsx"): fn += ".xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{fn}"')
                self.send_header("Content-Length", str(len(xlsx_bytes)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(xlsx_bytes)
            except Exception as ex:
                self.send_response(500)
                self.send_header("Content-Type","text/plain")
                self.send_header("Access-Control-Allow-Origin","*")
                self.end_headers()
                self.wfile.write(str(ex).encode())
        else:
            self.send_response(404); self.end_headers()

if __name__ == "__main__":
    print(f"Meeting Street Excel Server running on http://localhost:{PORT}")
    print("Keep this running while using the Time Tracker app.")
    print("Press Ctrl+C to stop.")
    with http.server.HTTPServer(("localhost", PORT), Handler) as srv:
        srv.serve_forever()
